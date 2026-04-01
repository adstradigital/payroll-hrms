import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal
from io import BytesIO

class ExcelExportService:
    @staticmethod
    def generate_salary_register(company, payroll_period, payslips):
        """
        Generates a detailed salary register where each column is a salary component.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Salary Register"

        # Define Styles
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        center_align = Alignment(horizontal="center", vertical="center")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # 1. Identify all unique salary components across these payslips
        earning_components = set()
        deduction_components = set()

        for payslip in payslips:
            for component in payslip.components.all():
                if component.component.component_type == 'earning':
                    earning_components.add(component.component.name)
                else:
                    deduction_components.add(component.component.name)

        earning_components = sorted(list(earning_components))
        deduction_components = sorted(list(deduction_components))

        # 2. Define Headers
        headers = ["Employee ID", "Name", "Department", "Designation", "Bank Name", "Account Number"]
        headers.extend(earning_components)
        headers.append("Gross Earnings")
        headers.extend(deduction_components)
        headers.append("Total Deductions")
        headers.append("Net Pay")

        # Write Headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border

        # 3. Write Data
        for row_num, payslip in enumerate(payslips, 2):
            emp = payslip.employee
            data = [
                emp.employee_id,
                emp.full_name,
                emp.department.name if emp.department else "N/A",
                emp.designation.name if emp.designation else "N/A",
                emp.bank_name or "N/A",
                emp.account_number or "N/A"
            ]

            # Map components to values
            comp_map = {c.component.name: c.amount for c in payslip.components.all()}
            
            # Add Earnings
            for comp_name in earning_components:
                data.append(comp_map.get(comp_name, 0))
            
            data.append(payslip.gross_earnings)

            # Add Deductions
            for comp_name in deduction_components:
                data.append(comp_map.get(comp_name, 0))
            
            data.append(payslip.total_deductions)
            data.append(payslip.net_salary)

            for col_num, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                if isinstance(value, (int, float, Decimal)):
                    cell.number_format = '#,##0.00'

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    @staticmethod
    def generate_payroll_summary(company, payroll_period, summary_data):
        """
        Generates a high-level summary report by Department.
        summary_data structure: [{'department': name, 'count': X, 'gross': Y, 'net': Z, 'epf': A, 'esi': B}]
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Payroll Summary"

        # Headers
        headers = ["Department", "Employee Count", "Gross Salary", "EPF Contribution", "ESI Contribution", "Net Payable"]
        
        # Style
        header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font

        # Data
        for row_num, item in enumerate(summary_data, 2):
            ws.cell(row=row_num, column=1, value=item['department'])
            ws.cell(row=row_num, column=2, value=item['count'])
            ws.cell(row=row_num, column=3, value=item['gross'])
            ws.cell(row=row_num, column=4, value=item['epf_employer'])
            ws.cell(row=row_num, column=5, value=item['esi_employer'])
            ws.cell(row=row_num, column=6, value=item['net'])

            # Formatting
            for col in range(3, 7):
                ws.cell(row=row_num, column=col).number_format = '#,##0.00'

        # Footer Totals
        last_row = len(summary_data) + 2
        ws.cell(row=last_row, column=1, value="TOTAL")
        ws.cell(row=last_row, column=1).font = Font(bold=True)
        
        for col in range(2, 7):
            col_letter = get_column_letter(col)
            ws.cell(row=last_row, column=col, value=f"=SUM({col_letter}2:{col_letter}{last_row-1})")
            ws.cell(row=last_row, column=col).font = Font(bold=True)
            ws.cell(row=last_row, column=col).number_format = '#,##0.00'

        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output
