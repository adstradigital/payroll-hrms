from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse


def get_date_range(period_type='month'):
    """
    Get start and end dates for common period types
    """
    today = datetime.now().date()
    
    if period_type == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = start_date + timedelta(days=6)
    elif period_type == 'month':
        start_date = today.replace(day=1)
        if today.month == 12:
            end_date = today.replace(day=31)
        else:
            end_date = (today.replace(month=today.month + 1, day=1) - timedelta(days=1))
    elif period_type == 'quarter':
        quarter = (today.month - 1) // 3
        start_date = today.replace(month=quarter * 3 + 1, day=1)
        end_date = (start_date + timedelta(days=90)).replace(day=1) - timedelta(days=1)
    elif period_type == 'year':
        start_date = today.replace(month=1, day=1)
        end_date = today.replace(month=12, day=31)
    else:
        start_date = today
        end_date = today
    
    return start_date, end_date


def calculate_days_remaining(target_date):
    """
    Calculate days remaining until target date
    """
    today = datetime.now().date()
    delta = target_date - today
    return delta.days


def format_rating(rating):
    """
    Format rating to 2 decimal places
    """
    if rating is None:
        return None
    return round(Decimal(str(rating)), 2)


def format_percentage(value):
    """
    Format value as percentage
    """
    if value is None:
        return "N/A"
    return f"{value:.2f}%"


def get_rating_color(rating, max_rating=5):
    """
    Get color code based on rating (for UI)
    Returns hex color code
    """
    if rating is None:
        return "#CCCCCC"  # Gray for no rating
    
    percentage = (rating / max_rating) * 100
    
    if percentage >= 80:
        return "#10B981"  # Green - Excellent
    elif percentage >= 60:
        return "#3B82F6"  # Blue - Good
    elif percentage >= 40:
        return "#F59E0B"  # Orange - Satisfactory
    else:
        return "#EF4444"  # Red - Needs Improvement


def export_reviews_to_excel(reviews, filename="performance_reviews.xlsx"):
    """
    Export performance reviews to Excel file
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Performance Reviews"
    
    # Define header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Employee Name', 'Employee ID', 'Department', 'Manager',
        'Review Period', 'Self Rating', 'Manager Rating', 'Overall Rating',
        'Rating Category', 'Status', 'Reviewed Date'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    for row_num, review in enumerate(reviews, 2):
        employee = review.employee
        
        worksheet.cell(row=row_num, column=1).value = employee.get_full_name()
        worksheet.cell(row=row_num, column=2).value = employee.id
        worksheet.cell(row=row_num, column=3).value = getattr(employee, 'department', 'N/A')
        worksheet.cell(row=row_num, column=4).value = review.reviewer.get_full_name() if review.reviewer else 'N/A'
        worksheet.cell(row=row_num, column=5).value = review.review_period.name
        worksheet.cell(row=row_num, column=6).value = float(review.self_rating) if review.self_rating else ''
        worksheet.cell(row=row_num, column=7).value = float(review.manager_rating) if review.manager_rating else ''
        worksheet.cell(row=row_num, column=8).value = float(review.overall_rating) if review.overall_rating else ''
        worksheet.cell(row=row_num, column=9).value = review.rating_category or ''
        worksheet.cell(row=row_num, column=10).value = review.get_status_display()
        worksheet.cell(row=row_num, column=11).value = review.reviewed_at.strftime('%Y-%m-%d') if review.reviewed_at else ''
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    workbook.save(response)
    return response


def export_bonus_projections_to_excel(projections, filename="bonus_projections.xlsx"):
    """
    Export bonus projections to Excel file
    """
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Bonus Projections"
    
    # Define header style
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    headers = [
        'Employee Name', 'Employee ID', 'Rating', 'Rating Category',
        'Base Salary', 'Bonus %', 'Bonus Amount'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    total_bonus = 0
    for row_num, proj in enumerate(projections, 2):
        worksheet.cell(row=row_num, column=1).value = proj['employee_name']
        worksheet.cell(row=row_num, column=2).value = proj['employee_id']
        worksheet.cell(row=row_num, column=3).value = proj['rating']
        worksheet.cell(row=row_num, column=4).value = proj['rating_category']
        worksheet.cell(row=row_num, column=5).value = proj['base_salary']
        worksheet.cell(row=row_num, column=6).value = proj['bonus_percentage']
        worksheet.cell(row=row_num, column=7).value = proj['bonus_amount']
        
        total_bonus += proj['bonus_amount']
    
    # Add total row
    total_row = len(projections) + 2
    worksheet.cell(row=total_row, column=6).value = "TOTAL:"
    worksheet.cell(row=total_row, column=6).font = Font(bold=True)
    worksheet.cell(row=total_row, column=7).value = total_bonus
    worksheet.cell(row=total_row, column=7).font = Font(bold=True)
    
    # Format currency columns
    for row in range(2, total_row + 1):
        worksheet.cell(row=row, column=5).number_format = '"$"#,##0.00'
        worksheet.cell(row=row, column=7).number_format = '"$"#,##0.00'
    
    # Auto-adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column = [cell for cell in column]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Create HTTP response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    workbook.save(response)
    return response


def get_performance_status_badge(status):
    """
    Get Bootstrap badge class for performance review status
    """
    status_badges = {
        'pending': 'badge-warning',
        'self_submitted': 'badge-info',
        'under_review': 'badge-primary',
        'completed': 'badge-success',
        'rejected': 'badge-danger',
    }
    return status_badges.get(status, 'badge-secondary')


def calculate_completion_percentage(completed, total):
    """
    Calculate completion percentage
    """
    if total == 0:
        return 0
    return round((completed / total) * 100, 2)
